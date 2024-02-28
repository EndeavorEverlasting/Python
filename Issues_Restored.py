import re, os, sys, logging, random
from shutil import copy2
#from typing import Union
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from collections import Counter


class DataProcessor:

  def __init__(self):
    self.input_file = "LIJ 2_2_24.xlsx"
    self.output_file = None
    self.row_counter = []

  def load_Data(self):
    # This function loads data from the excel sheet and adds a column for issues
    # The "Issues" column will be the first column (column 1)

    self.wb = load_workbook(self.input_file, read_only=False, data_only=False)
    self.ws = self.wb.active  # This part picks the first sheet on the excel

    # Call update_headers() to remove "Department_ID" and update other headers
    self.update_headers()

    # Add the "Issues" column at the same column position as "Department_ID"
    self.issueColumn = 6  # Column F
    self.ws.cell(row=2, column=self.issueColumn).value = "Issues"

    self.initialize_column_index_map()  # Run after adding column for issues
    for c in self.ws["CD"]:
      c.value = ""
    return None

  def update_headers(self):
    # Remove the "Department_ID" column
    #for row in self.ws.iter_rows(min_row=1, max_row=1):
    #  for cell in row:
    #    if cell.value == "Department_ID":
    #      self.ws.delete_cols(cell.column, 1)
    #      break

    # Update other headers and column index map
    header_row = self.ws[1]
    column_names = [cell.value for cell in header_row]
    self.column_index_map = {
        name: index + 1
        for index, name in enumerate(column_names)
    }

  def initialize_column_index_map(self):
    # This function initializes a dictionary to map column names to column indices
    header_row = self.ws[2]
    column_names = [cell.value for cell in header_row]
    self.column_index_map = {
        name: index + 1
        for index, name in enumerate(column_names)
    }

    print("Header Row: ", header_row)
    print("Column Index Map:", self.column_index_map)

  def setup_logging(self):
    #this function sets up logging, and sets up the directory to log data to.
    output_folder = 'Outputs/'
    output_directory = os.path.join(output_folder, 'logs')
    os.makedirs(output_directory, exist_ok=True)
    log_file = os.path.join('Outputs', 'logs', 'data_processor.log')
    logging.basicConfig(
        filename=log_file,  # Set the log file name
        level=logging.INFO,  # Set the logging level to INFO
        format='%(asctime)s - %(levelname)s - %(message)s')  # format 3

  def highlight(self, cell, color):
    #this function highlihgts the cell bright purple and the row yellow - red
    #if the row shows up 1 for issue make yellow 2x orange 3+ red and highlight purple issue cell
    yellow = "00FFFF66"
    orange = "00FF9933"
    red = "00FF0000"
    row = cell.row
    if self.row_counter.count(row) <= 1:
      for i in range(1,82):
        x = self.ws.cell(row=row, column=i)
        x.fill = PatternFill(start_color=yellow,end_color=yellow,fill_type="solid")
    elif self.row_counter.count(row) == 2:
      for i in range(1,82):
        x = self.ws.cell(row=row, column=i)
        x.fill = PatternFill(start_color=orange,end_color=orange,fill_type="solid")
    elif self.row_counter.count(row == 3):
      for i in range(1,82):
        x = self.ws.cell(row=row, column=i)
        x.fill = PatternFill(start_color=red,end_color=red,fill_type="solid")
    else:
      pass
    cell.fill = PatternFill(start_color=color,
                            end_color=color,
                            fill_type="solid")
    self.row_counter.append(row)
    return None

  def validate_floor_plans(self):
    # Extract all floor plans and designators from the worksheet
    floor_plans_designators = [
        (str(
            self.ws.cell(
                row=i,
                column=self.column_index_map["Flr Pln N"]).value).strip(),
         str(
             self.ws.cell(
                 row=i,
                 column=self.column_index_map["Flr Pln D"]).value).strip())
        for i in range(3, self.ws.max_row + 1)
    ]

    # Initialize a dictionary to store designators for each floor plan
    floor_plan_designators_map = {}

    # Group designators by floor plan
    for floor_plan, designator in floor_plans_designators:
      floor_plan_designators_map.setdefault(floor_plan, []).append(designator)

    # Sort the designators within each floor plan group
    for floor_plan in floor_plan_designators_map:
      floor_plan_designators_map[floor_plan] = sorted(
          floor_plan_designators_map[floor_plan])

    return floor_plan_designators_map

  def floorPlanIssues(self, data):
    #this function checks for missing floor plans
    #floor data[1] and data[2] are the old and new floor plans respectively
    #they are located at K and L on excel sheet

    if data[1] == None and data[2] == None:
      self.highlight(
          self.ws.cell(row=data[0], column=self.column_index_map["Flr Pln L"]),"00FF00FF")
      self.highlight(
          self.ws.cell(row=data[0], column=self.column_index_map["Flr Pln N"]),"00FF00FF")
      return "/missing a floor plan"
    return None

  def designatorIssues(self, data):
    #this function checks for missing designators
    #they are located at M on the excel sheet
    if data[1] is None:
      self.highlight(
          self.ws.cell(row=data[0], column=self.column_index_map["Flr Pln D"]),"00FF00FF")
      return "/needs a designator"
    return ""

  def find_monitor_issues(self, data):
    #this function checks for missing monitors
    #3 things can be missing on monitor info at Y,Z,AA
    if self.ws[f"B{data[0]}"].value.lower() in "workstation,laptop,desktop":
      if data[1] == None:
        self.ws.cell(row=data[0], column=self.issueColumn).value = str(
            self.ws.cell(
                row=data[0],
                column=self.issueColumn).value) + "/needs a monitor make"
      if data[2] == None:
        self.ws.cell(row=data[0], column=self.issueColumn).value = str(
            self.ws.cell(
                row=data[0],
                column=self.issueColumn).value) + "/needs a monitor model"
      if data[3] == None:
        self.ws.cell(row=data[0], column=self.issueColumn).value = str(
            self.ws.cell(
                row=data[0],
                column=self.issueColumn).value) + "/needs a monitor size"
    return None

  def check_duplicate_designators(self):
    duplicates = {}
    for i in range(3, self.ws.max_row + 1):
      # Extract the shared portion of the floor plan name
      flr_pln_n = str(
          self.ws.cell(
              row=i, column=self.column_index_map["Flr Pln N"]).value).strip()
      flr_pln_d = str(
          self.ws.cell(
              row=i, column=self.column_index_map["Flr Pln D"]).value).strip()
      if flr_pln_n and flr_pln_d:  # Ignore blanks
        # Extract the shared portion of the floor plan name
        floor_plan_shared = re.search(r'^LIJMC - \d+', flr_pln_n)
        if floor_plan_shared:
          floor_plan_shared = floor_plan_shared.group()
        else:
          floor_plan_shared = ""

        condition = (flr_pln_d, floor_plan_shared)

        # Check if the condition is already in the dictionary
        # Mark conditions by row for spreed sheet
        if condition in duplicates:
          duplicates[condition].append(i)
        else:
          duplicates[condition] = [i]
    for condition, rows in duplicates.items():
      if len(rows) > 1:
        self.ws.cell(row=rows[0], column=self.issueColumn).value = (
            f"Duplicates with {', '.join(str(row) for row in rows if row != rows[0])}"
        )

  def sequence_check(self, row_id, flr_pln_L_value, flr_pln_N_value,
                     department_value, flr_pln_D_value):
    try:
      pass
      #print("sequence_check() start")

      #print("sequence_check() validate floor plans")

      # Approach 3: Generate a list of designators for each valid floor plan
      #valid_floor_plans = self.validate_floor_plans()

      # Print the count of each floor plan name for debugging
      #print("Count of each floor plan name:")
      #for floor_plan, designators in valid_floor_plans.items():
      #print(f"{floor_plan}: {len(designators)}")

      # Check the sequence for each group of designators
      # associated with a floor plan
      #if not self.are_designators_sequential(designators, floor_plan):
      # Apply light yellow highlight to the cell with the issue
      #  floor_plan_cell = self.ws.cell(
      #      row=row_id, column=self.column_index_map["Flr Pln D"])
      #  print("sequence_check() highlighting the cell")
      print("sequence_check() finish")

    except Exception as e:
      print(
          f"Error in sequence_check: {e}\nLine: {sys.exc_info()[-1].tb_lineno}"
      )
    return None

  import re

  def are_designators_sequential(self, designators, floor_plan_name):
    """
      Check if the given list of designators is sequential based on the floor plan name.
      """
    designator_numbers = []
    for designator in designators:
      match = re.search(r'\d+', designator)
      if match:
        designator_numbers.append(int(match.group()))
      else:
        print(
            f"No numerical part found for designator: {designator} in floor plan {floor_plan_name}"
        )

    # Sort the designator numbers
    sorted_numbers = sorted(designator_numbers)

    # Check if the sorted numbers form a sequential sequence
    return sorted_numbers == list(
        range(min(sorted_numbers),
              max(sorted_numbers) + 1))

  def printer_Issues(self, data):
    # This function looks for printer errors
    # data1 = bc data2 = bf data3 = bd data4 = bg data5 = bh
    # bc = type bf = ip bd = queue name bg = make bh = model
    issue_message = ""
    if data[1] is None:
      issue_message += "/Missing printer Type"
    if data[2] is None:
      issue_message += "/Missing printer IP"
    if data[3] is None:
      issue_message += "/Missing printer Queue Name"
    if data[4] is None:
      issue_message += "/Missing printer Make"
    if data[5] is None:
      issue_message += "/Missing printer Model"

    if issue_message:
      self.ws.cell(
          row=data[0], column=self.issueColumn
      ).value += issue_message  # Convert to string before concatenation
    else:
      self.ws.cell(
          row=data[0], column=self.issueColumn
      ).value = issue_message  # Convert to string before assignment
    return None

  def find_duplicates(self):
    # Initialize a dictionary to store duplicate records
    duplicates = {}

    # Iterate over each row in the worksheet
    for i in range(3, self.ws.max_row + 1):
      # Extract the necessary data for identifying duplicates
      flr_pln_d = self.ws.cell(row=i,
                               column=self.column_index_map["Flr Pln D"]).value
      flr_pln_n = self.ws.cell(row=i,
                               column=self.column_index_map["Flr Pln N"]).value
      department = self.ws.cell(
          row=i, column=self.column_index_map["Department"]).value

      # Define a condition based on the extracted data
      condition = (flr_pln_d, flr_pln_n, department)

      # Check if the condition already exists in the duplicates dictionary
      if condition in duplicates:
        duplicates[condition].append(i)
      else:
        duplicates[condition] = [i]

    # Return the dictionary containing duplicate records
    return duplicates

#
#
#
#
#

  def check_for_duplicates(self):
    # Step 1: Call the find_duplicates method
    duplicates = self.find_duplicates()

    # Step 2: Flag issues
    issue_messages = []
    for i in range(3, self.ws.max_row + 1):
      # Step 2a: Collect all issue messages for the current row

      # Step 2b: Check for floor plan and duplicate issues
      issue_messages.append(
          self.floorPlanIssues(
              [i, self.ws[f"K{i}"].value, self.ws[f"L{i}"].value]))
      issue_messages.append(self.designatorIssues([i, self.ws[f"M{i}"].value]))

      # Collect issue messages from other functions as well

      # Step 2c: Check for monitor issues:
      # checks for issues with monitor designator and floorplans
      for i in range(3, self.ws.max_row + 1):
        self.floorPlanIssues(
            [i, self.ws[f"K{i}"].value, self.ws[f"L{i}"].value])
        self.designatorIssues([i, self.ws[f"M{i}"].value])
        # list may need to extend in future because monitor goes up to 4
        self.find_monitor_issues([
            i, self.ws[f"Y{i}"].value, self.ws[f"Z{i}"].value,
            self.ws[f"AA{i}"].value
        ])

    # Step 2: collecting for sequence_check:
    flr_pln_L_value = self.ws.cell(
        row=i, column=self.column_index_map["Flr Pln L"]).value
    flr_pln_N_value = self.ws.cell(
        row=i, column=self.column_index_map["Flr Pln N"]).value
    department_value = self.ws.cell(
        row=i, column=self.column_index_map["Department"]).value
    flr_pln_D_value = self.ws.cell(
        row=i, column=self.column_index_map["Flr Pln D"]).value

    # Step 2a: Check for sequence issues on data:
    self.sequence_check(i, flr_pln_L_value, flr_pln_N_value, department_value,
                        flr_pln_D_value)

    # Step 3: Combine issue messages into one string
    for i in range(3, self.ws.max_row + 1):
      combined_message = " ".join(msg for msg in issue_messages if msg)
      self.ws.cell(row=i, column=self.issueColumn).value = combined_message

    return None

  def flagging_issues(self):
    # Step 1: Call the find_duplicates method
    duplicates = self.find_duplicates()

    # Step 2: Flag issues
    issue_messages = []
    for i in range(3, self.ws.max_row + 1):

      # Step 3: Collect all issue messages for the current row
      self.ws.cell(row=i, column=self.issueColumn).value = ""
      row_issues = []

      # Step 4: Check for floorplan issues
      row_issues.append(
          self.floorPlanIssues(
              [i, self.ws[f"K{i}"].value, self.ws[f"L{i}"].value]))

      # Step 5: Check for designator issues
      row_issues.append(self.designatorIssues([i, self.ws[f"M{i}"].value]))

      #Step 6: Check for monitor issues
      self.find_monitor_issues([
          i, self.ws[f"Y{i}"].value, self.ws[f"Z{i}"].value,
          self.ws[f"AA{i}"].value
      ])

      # Step 7: Check for printer issues
      if "printer" in self.ws.cell(
          row=i, column=self.column_index_map["Type"]).value.lower():
        self.printer_Issues([
            i,
            self.ws.cell(row=i,
                         column=self.column_index_map["PRNT_Type"]).value,
            self.ws.cell(
                row=i, column=self.column_index_map["Network Pntr IP"]).value,
            self.ws.cell(
                row=i, column=self.column_index_map["PRNT_Queue_Name"]).value,
            self.ws.cell(row=i,
                         column=self.column_index_map["PRNT_Make"]).value,
            self.ws.cell(row=i,
                         column=self.column_index_map["PRNT_Model"]).value
        ])

      # Step 8: Check for sequence issues on data:
      self.sequence_check(
          i,
          self.ws.cell(row=i, column=self.column_index_map["Flr Pln L"]).value,
          self.ws.cell(row=i, column=self.column_index_map["Flr Pln N"]).value,
          self.ws.cell(row=i,
                       column=self.column_index_map["Department"]).value,
          self.ws.cell(row=i, column=self.column_index_map["Flr Pln D"]).value)

      # Collect issue messages for the current row
      issue_messages.append(row_issues)

    # Step 9: Combine issue messages into one string for each row
    for i in range(3, self.ws.max_row + 1):
      combined_message = " ".join(msg for msg in issue_messages[i - 3] if msg)
      self.ws.cell(row=i, column=self.issueColumn).value = combined_message

    return None

  def count_devices(self):
    # Initialize counts for laptops, WOWs, workstations, printers, and specialty printers
    laptop_count = 0
    wow_count = 0
    workstation_count = 0
    printer_count = 0
    specialty_printer_count = 0

    # Iterate over each row in the worksheet
    for i in range(3, self.ws.max_row + 1):
      # Check if the type is 'Workstation' or 'Printer'
      device_type = self.ws.cell(
          row=i, column=self.column_index_map["Type"]).value.lower()
      if device_type == 'workstation':
        # Check if the designator starts with
        # 'W' for WOWs, 'L' for Laptops, or 'W{digits}' for Workstations
        designator = str(
            self.ws.cell(
                row=i,
                column=self.column_index_map["Flr Pln D"]).value).strip()
        if designator:
          if designator[0].lower() == 'l':
            laptop_count += 1
          elif designator.startswith('wow', 0, 3):
            wow_count += 1
          elif designator[0].lower() == 'w':
            workstation_count += 1
      elif device_type == 'printer':
        # Check if the designator starts with 'P' for Printers or 'S' for Specialty Printers
        designator = str(
            self.ws.cell(
                row=i,
                column=self.column_index_map["Flr Pln D"]).value).strip()
        if designator:
          if designator[0].lower() == 'p':
            printer_count += 1
          elif designator[0].lower() == 's':
            specialty_printer_count += 1

    # Return the counts
    return {
        'Laptops': laptop_count,
        'WOWs': wow_count,
        'Workstations': workstation_count,
        'Printers': printer_count,
        'Specialty Printers': specialty_printer_count
    }

  def save_output_file(self):
    print(f"Beginning the determine_output_files(): {self.output_file}")
    output_folder = 'Outputs/'
    current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_extension = os.path.splitext(self.input_file)[1].lower()
    file_type = {
        'csv': 'csv',
        'xlsx': 'ss',
        'txt': 'txt',
        'log': 'logs'
    }.get(file_extension, 'ss')

    output_file_name = f"output_{current_datetime}.xlsx"
    self.output_file = os.path.join(output_folder, file_type, output_file_name)

    # Check if the output_file already exists
    if os.path.exists(self.output_file):
      # Add a copy number if the file already exists
      copy_number = 1
      while os.path.exists(self.output_file):
        output_file_name = f"output_{current_datetime}_copy{copy_number}.xlsx"
        self.output_file = os.path.join(output_folder, file_type,
                                        output_file_name)
        copy_number += 1

    self.wb.save(self.output_file)
    print(f"Ending the save_output_file(): {self.output_file}")
    print(self.output_file)
    return None

  def process_data(self):
    print(f"Beginning of process_data: {self.output_file}")
    self.setup_logging()
    self.load_Data()
    self.flagging_issues()
    self.save_output_file()
    print(f"Ending of process_data: {self.output_file}")
    return None


# Instantiate the DataProcessor class
print("Starting the data the processor")
data_processor = DataProcessor()

# Call the process_data method on the instance
data_processor.process_data()

print("After calling the process_data method on the instance")
# Configure logging to write messages/logs to a file
logging.basicConfig(filename='data_processing.log', level=print)
